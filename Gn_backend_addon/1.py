import openpyxl
n = str(9) 
xlsx = 'txsedb/T'+n+'SE.xlsx'


wb = openpyxl.load_workbook(xlsx)
ws = wb.active
max_row = ws.max_row


raw_total_fasta =  'T'+n+'SE.fasta'

total_outfile = open(raw_total_fasta,'w',encoding='utf-8')
for  i in range(2,max_row+1):
    k = ws.cell(row = i ,column = 12).value
    protein_id = ws.cell(row=i, column=5).value
    sequence = ws.cell(row=i, column=14).value
    total_outfile.write(">"+protein_id+'\n'+sequence+'\n')
    


total_outfile.close()
wb.save(xlsx)